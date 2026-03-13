import pymorphy3
import re
from collections import defaultdict, Counter
from typing import Dict
import json
import openpyxl
from openpyxl.styles import Font
from io import BytesIO


def find_word_form(content: str) -> Dict:
    """Возвращает статистику словоформ (JSON)"""
    morph = pymorphy3.MorphAnalyzer()
    lines = content.splitlines()
    num_lines = len(lines)
    stats = defaultdict(lambda: {'total': 0, 'line_counts': []})

    for line_num, line in enumerate(lines):
        words = re.findall(r'[а-яё]+', line.lower())
        line_lemmas = []
        for word in words:
            if len(word) > 2:
                parsed = morph.parse(word)[0]
                lemma = parsed.normal_form
                if lemma and len(lemma) > 2:
                    line_lemmas.append(lemma)

        line_counter = Counter(line_lemmas)
        for lemma in line_counter:
            stats[lemma]['total'] += line_counter[lemma]
            stats[lemma]['line_counts'].append(line_counter[lemma])

    with open('app/API/endpoints/data.json', 'w', encoding='utf-8') as f:
        json.dump({
            'stats': dict(stats),
            'num_lines': num_lines
        }, f, ensure_ascii=False, indent=2)

    return {
        "status": "processed",
        "word_forms_count": len(stats),
        "total_words": sum(s['total'] for s in stats.values()),
        "message": "Данные сохранены в data.json."
    }


def create_excel_from_stats(stats_file: str = 'app/API/endpoints/data.json') -> BytesIO:
    """Дорабатывает JSON в Excel"""
    with open(stats_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    stats = data['stats']
    num_lines = data['num_lines']
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Частотная статистика"
    ws['A1'] = 'Словоформа'
    ws['B1'] = 'Общее количество'
    ws['C1'] = 'По строкам'

    for col in ['A1', 'B1', 'C1']:
        ws[col].font = Font(bold=True)
    sorted_stats = sorted(stats.items(), key=lambda x: x[1]['total'], reverse=True)

    for row_idx, (lemma, data) in enumerate(sorted_stats, 2):
        total_count = data['total']
        line_counts = data['line_counts']
        full_line_counts = line_counts + [0] * (num_lines - len(line_counts))
        line_counts_str = ','.join(map(str, full_line_counts))
        ws[f'A{row_idx}'] = lemma
        ws[f'B{row_idx}'] = total_count
        ws[f'C{row_idx}'] = line_counts_str

    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    wb.save(output)
    output.seek(0)
    return output
